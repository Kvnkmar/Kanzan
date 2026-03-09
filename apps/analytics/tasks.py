"""
Celery tasks for the analytics app.

Provides asynchronous processing of data export jobs, generating CSV or XLSX
files and updating job status on completion or failure.
"""

import csv
import io
import logging
from datetime import datetime

from celery import shared_task
from django.core.files.base import ContentFile
from django.utils import timezone

logger = logging.getLogger(__name__)


@shared_task(bind=True, max_retries=3, default_retry_delay=60)
def process_export_job(self, export_job_id):
    """
    Process a pending ExportJob by generating the requested file.

    Steps:
        1. Mark the job as ``processing``.
        2. Fetch the appropriate queryset based on ``resource_type`` and ``filters``.
        3. Generate a CSV or XLSX file.
        4. Attach the file to the ExportJob and mark it as ``completed``.
        5. On failure, record the error and mark the job as ``failed``.
    """
    from apps.analytics.models import ExportJob

    try:
        job = ExportJob.unscoped.select_related("tenant", "report").get(
            id=export_job_id
        )
    except ExportJob.DoesNotExist:
        logger.error("ExportJob %s not found.", export_job_id)
        return

    # Mark as processing.
    job.status = ExportJob.Status.PROCESSING
    job.save(update_fields=["status", "updated_at"])

    try:
        rows, headers = _fetch_export_data(job)
        file_content, filename = _generate_file(job, headers, rows)

        job.file.save(filename, ContentFile(file_content), save=False)
        job.status = ExportJob.Status.COMPLETED
        job.completed_at = timezone.now()
        job.save(update_fields=["file", "status", "completed_at", "updated_at"])

        logger.info(
            "ExportJob %s completed: %s (%d rows).",
            export_job_id,
            filename,
            len(rows),
        )

    except Exception as exc:
        logger.exception("ExportJob %s failed.", export_job_id)
        job.status = ExportJob.Status.FAILED
        job.error_message = str(exc)
        job.save(update_fields=["status", "error_message", "updated_at"])
        raise self.retry(exc=exc)


def _fetch_export_data(job):
    """
    Fetch data rows and headers based on the export job's resource_type and filters.

    Returns:
        (rows, headers) where rows is a list of dicts and headers is a list of
        column names.
    """
    tenant = job.tenant
    filters = job.filters or {}

    if job.resource_type == "tickets":
        return _fetch_ticket_data(tenant, filters)
    elif job.resource_type == "contacts":
        return _fetch_contact_data(tenant, filters)
    else:
        raise ValueError(f"Unsupported resource type: {job.resource_type}")


def _fetch_ticket_data(tenant, filters):
    """Fetch ticket data for export."""
    from apps.tickets.models import Ticket

    qs = Ticket.unscoped.filter(tenant=tenant).select_related(
        "status", "assignee", "queue", "contact"
    )

    if filters.get("status"):
        qs = qs.filter(status__slug=filters["status"])
    if filters.get("priority"):
        qs = qs.filter(priority=filters["priority"])
    if filters.get("assignee"):
        qs = qs.filter(assignee_id=filters["assignee"])
    if filters.get("date_from"):
        qs = qs.filter(created_at__gte=filters["date_from"])
    if filters.get("date_to"):
        qs = qs.filter(created_at__lte=filters["date_to"])

    headers = [
        "number",
        "subject",
        "status",
        "priority",
        "assignee",
        "queue",
        "contact",
        "created_at",
        "resolved_at",
    ]

    rows = []
    for ticket in qs.iterator():
        rows.append(
            {
                "number": ticket.number,
                "subject": ticket.subject,
                "status": ticket.status.name if ticket.status else "",
                "priority": ticket.get_priority_display(),
                "assignee": (
                    ticket.assignee.get_full_name() if ticket.assignee else ""
                ),
                "queue": ticket.queue.name if ticket.queue else "",
                "contact": str(ticket.contact) if ticket.contact else "",
                "created_at": (
                    ticket.created_at.isoformat() if ticket.created_at else ""
                ),
                "resolved_at": (
                    ticket.resolved_at.isoformat() if ticket.resolved_at else ""
                ),
            }
        )

    return rows, headers


def _fetch_contact_data(tenant, filters):
    """Fetch contact data for export."""
    from apps.contacts.models import Contact

    qs = Contact.unscoped.filter(tenant=tenant).select_related("company")

    if filters.get("source"):
        qs = qs.filter(source=filters["source"])
    if filters.get("is_active") is not None:
        qs = qs.filter(is_active=filters["is_active"])
    if filters.get("company"):
        qs = qs.filter(company_id=filters["company"])
    if filters.get("date_from"):
        qs = qs.filter(created_at__gte=filters["date_from"])
    if filters.get("date_to"):
        qs = qs.filter(created_at__lte=filters["date_to"])

    headers = [
        "first_name",
        "last_name",
        "email",
        "phone",
        "company",
        "job_title",
        "source",
        "is_active",
        "created_at",
    ]

    rows = []
    for contact in qs.iterator():
        rows.append(
            {
                "first_name": contact.first_name,
                "last_name": contact.last_name,
                "email": contact.email,
                "phone": contact.phone or "",
                "company": contact.company.name if contact.company else "",
                "job_title": contact.job_title or "",
                "source": contact.get_source_display() if contact.source else "",
                "is_active": contact.is_active,
                "created_at": (
                    contact.created_at.isoformat() if contact.created_at else ""
                ),
            }
        )

    return rows, headers


def _generate_file(job, headers, rows):
    """
    Generate the export file in the requested format.

    Returns:
        (file_bytes, filename)
    """
    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"{job.resource_type}_{timestamp}"

    if job.export_type == "csv":
        return _generate_csv(headers, rows), f"{base_name}.csv"
    elif job.export_type == "xlsx":
        return _generate_xlsx(headers, rows), f"{base_name}.xlsx"
    elif job.export_type == "pdf":
        # PDF generation is a placeholder; production would use a library
        # like ReportLab or WeasyPrint.
        return _generate_csv(headers, rows), f"{base_name}.csv"
    else:
        raise ValueError(f"Unsupported export type: {job.export_type}")


def _generate_csv(headers, rows):
    """Generate a CSV file as bytes."""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return output.getvalue().encode("utf-8")


def _generate_xlsx(headers, rows):
    """
    Generate an XLSX file as bytes.

    Falls back to CSV if ``openpyxl`` is not installed.
    """
    try:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export"

        # Write header row.
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # Write data rows.
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    except ImportError:
        logger.warning(
            "openpyxl is not installed; falling back to CSV for XLSX export."
        )
        return _generate_csv(headers, rows)
